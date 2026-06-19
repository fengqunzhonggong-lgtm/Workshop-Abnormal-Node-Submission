import io
from datetime import datetime
from flask import Blueprint, request, jsonify, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from database import get_db
from auth import require_auth, require_role

export_bp = Blueprint('export', __name__)


@export_bp.route('/api/export/excel', methods=['GET'])
@require_auth
@require_role('manager', 'superadmin')
def export_excel():
    where, params = [], []
    for f, op in [('date_from', '>='), ('date_to', '<=')]:
        v = request.args.get(f)
        if v:
            where.append(f"date(r.created_at) {op} ?")
            params.append(v)
    for f in ['work_order_no', 'product_model_id', 'abnormal_type_id', 'source_department_id']:
        v = request.args.get(f)
        if v:
            where.append(f"r.{f} = ?")
            params.append(v)
    product_model_name = request.args.get('product_model_name')
    if product_model_name:
        where.append("pm.name LIKE ?")
        params.append(f"%{product_model_name}%")
    where_clause = ('WHERE ' + ' AND '.join(where)) if where else ''

    db = get_db()
    rows = db.execute(f"""
        SELECT r.*,
            pm.name as product_model_name,
            at2.name as abnormal_type_name,
            sd.name as source_department_name,
            sp.name as source_process_name,
            fp.name as found_process_name
        FROM abnormal_records r
        LEFT JOIN product_models pm ON r.product_model_id = pm.id
        LEFT JOIN abnormal_types at2 ON r.abnormal_type_id = at2.id
        LEFT JOIN source_departments sd ON r.source_department_id = sd.id
        LEFT JOIN source_processes sp ON r.source_process_id = sp.id
        LEFT JOIN found_processes fp ON r.found_process_id = fp.id
        {where_clause}
        ORDER BY r.created_at DESC
    """, params).fetchall()
    db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = '异常明细'

    headers = ['序号', '异常编号', '工单号', '产品型号', '异常类别', '数量',
               '来源部门', '来源工序', '发现工序', '异常描述', '提交人', '工号', '提交时间']
    ws.append(headers)

    for i, r in enumerate(rows, 1):
        ws.append([
            i, r['record_no'], r['work_order_no'], r['product_model_name'],
            r['abnormal_type_name'], r['quantity'], r['source_department_name'],
            r['source_process_name'], r['found_process_name'], r['description'],
            r['submitter_name'], r['submitter_employee_id'], r['created_at']
        ])

    # Style headers
    header_style = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ts = datetime.now().strftime('%Y%m%d')
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="anomaly_export_{ts}.xlsx"'}
    )
